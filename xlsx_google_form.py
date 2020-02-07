#Author: Jack Schwarz
#Date: 1/20/2020
#Version: 1.0

import openpyxl
import requests

form_ID = "https://docs.google.com/forms/d/e/1FAIpQLScItCDe-NY55Tb8Zhjh9YXzKrN6CxG1JfUxN0zNsr0Kc88kqQ/formResponse" #Form Response URL to Google Form
file_path = ("/Users/jack/Documents/CoboQuestions.xlsx") #File path from root directory to .xlsx document to be parsed
person_name = "Jack Schwarz"

wbook = openpyxl.load_workbook(file_path, data_only=True) #Opens .xlsx document for readin/writing
sheet = wbook.active

#The following are for my personal use. Pulls values out of .xlsx to keep track of how many cells have been parsed
number_of_questions = sheet.cell(row = 2, column = 6).value
last_submitted_question = sheet.cell(row = 4, column = 6).value
initial_question = last_submitted_question

number_of_rows = sheet.max_row #Row number to determine how many to parse

if(number_of_questions != last_submitted_question): #Making sure we have data that hasn't been read
    for i in range(number_of_questions - last_submitted_question):
        last_submitted_question += 1 #openpyxl starts rows at 1
        last_submitted_question = sheet.cell(row = last_submitted_question + 1, column = 1).value 

        #Used to create submission (cleaner than all in one line)
        qNum = i 
        qNum += 1
        qSpin = sheet.cell(row = last_submitted_question + 1, column = 4).value
        qText = sheet.cell(row = last_submitted_question + 1, column = 2).value
        qAnsw = sheet.cell(row = last_submitted_question + 1, column = 3).value

        submission = {'entry.1229841533': person_name, 'entry.544582793': qNum, 'entry.1388940012': qSpin, 'entry.1456041777': qText, 'entry.380309657': qAnsw}

        response = requests.post(form_ID, submission) #Sends out request to Form

        print(str(qNum) + ' submitted: ' + str(qSpin) + ' --- ' + str(qText) + ' --- '+ str(qAnsw)) #Prints to confirm it has been parsed

print(last_submitted_question)

#Changing indexes in the spreadsheet
sheet['F4'] = last_submitted_question
sheet['F4'].value
sheet['F2'] = '=COUNTA(B2:B101)' #Runs excel 'COUNTA' formula

wbook.save(file_path) #Saves spreadsheet
